from __future__ import annotations

import argparse
import csv
import json
import re
from collections.abc import Iterable, Sequence
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]
from sqlalchemy.orm import Session

from app.core.database import SessionLocal
from app.repositories.people import PersonRepository, PlayerRepository
from app.repositories.positions import PositionRepository
from app.repositories.rosters import PlayerRosterRepository
from app.repositories.schools import CollegeRepository, HighSchoolRepository
from app.schemas.roster_import import RosterImportRow, RosterImportSummary
from app.services.imports import RosterImportService
from app.utils.roster_normalization import clean_text

DEFAULT_SHEET_NAME = "Cleaned Data"

COLUMN_ALIASES: dict[str, set[str]] = {
    "college": {"college", "school", "team", "university", "institution"},
    "college_state": {"state", "college_state", "school_state"},
    "college_division": {"division", "div"},
    "conference": {"conference", "conf"},
    "full_name": {"name", "player", "player_name", "full_name", "student_athlete"},
    "first_name": {"first_name", "firstname"},
    "last_name": {"last_name", "lastname"},
    "roster_year": {"year", "yr", "class", "roster_year", "academic_year"},
    "position": {"position", "pos", "positions"},
    "height": {"height", "ht"},
    "weight": {"weight", "wt"},
    "bats_throws": {"b_t", "bt", "bats_throws", "bats_throws_combined"},
    "bats": {"bats", "bat"},
    "throws": {"throws", "throw"},
    "hometown": {"hometown", "home_town"},
    "hometown_city": {"hometown_city", "home_city", "city"},
    "hometown_state": {"hometown_state", "home_state"},
    "hometown_country": {"hometown_country", "home_country", "country"},
    "high_school": {"high_school", "hs", "prep_school"},
    "previous_school": {"previous_school", "previous_college", "prev_school", "transfer_from"},
    "is_transfer": {"transfer", "is_transfer", "transfer_flag"},
    "jersey_number": {"number", "no", "num", "jersey", "jersey_number"},
    "roster_url": {"roster_url", "source_url"},
    "profile_url": {"profile_url", "bio_url", "player_url", "url"},
}


def _normalize_header(value: object) -> str:
    text = clean_text(value)
    if text is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "_", text.casefold()).strip("_")


ALIAS_TO_FIELD = {
    _normalize_header(alias): field_name
    for field_name, aliases in COLUMN_ALIASES.items()
    for alias in aliases
}


def _cell_to_text(value: object) -> str | None:
    return clean_text(value)


def read_roster_rows(path: Path, sheet_name: str = DEFAULT_SHEET_NAME) -> list[RosterImportRow]:
    suffix = path.suffix.casefold()
    if suffix == ".xlsx":
        return read_xlsx_roster_rows(path, sheet_name)
    if suffix == ".csv":
        return read_csv_roster_rows(path)
    raise ValueError(f"unsupported roster import file type: {path.suffix}")


def read_xlsx_roster_rows(
    path: Path, sheet_name: str = DEFAULT_SHEET_NAME
) -> list[RosterImportRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"worksheet not found: {sheet_name}")
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers is None:
            return []
        return _rows_from_values(headers, rows, first_data_row_number=2)
    finally:
        workbook.close()


def read_csv_roster_rows(path: Path) -> list[RosterImportRow]:
    with path.open("r", encoding="utf-8-sig", newline="") as roster_file:
        reader = csv.reader(roster_file)
        headers = next(reader, None)
        if headers is None:
            return []
        return _rows_from_values(headers, reader, first_data_row_number=2)


def _rows_from_values(
    headers: Sequence[object], rows: Iterable[Sequence[object]], first_data_row_number: int
) -> list[RosterImportRow]:
    field_names = [_field_name_for_header(header) for header in headers]
    import_rows: list[RosterImportRow] = []
    for offset, values in enumerate(rows):
        row_data: dict[str, str | int | None] = {"row_number": first_data_row_number + offset}
        has_value = False
        for field_name, value in zip(field_names, values, strict=False):
            if field_name is None:
                continue
            text = _cell_to_text(value)
            if text is None:
                continue
            has_value = True
            row_data.setdefault(field_name, text)
        if has_value:
            import_rows.append(RosterImportRow.model_validate(row_data))
    return import_rows


def _field_name_for_header(header: object) -> str | None:
    normalized_header = _normalize_header(header)
    return ALIAS_TO_FIELD.get(normalized_header)


def build_service(session: Session) -> RosterImportService:
    return RosterImportService(
        colleges=CollegeRepository(session),
        people=PersonRepository(session),
        players=PlayerRepository(session),
        positions=PositionRepository(session),
        high_schools=HighSchoolRepository(session),
        rosters=PlayerRosterRepository(session),
    )


def run_import(path: Path, year: int, sheet_name: str = DEFAULT_SHEET_NAME) -> RosterImportSummary:
    rows = read_roster_rows(path, sheet_name)
    session = SessionLocal()
    try:
        service = build_service(session)
        summary = service.import_rows(rows, year)
        session.commit()
        return summary
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Import roster rows from a cleaned NCAA roster file."
    )
    parser.add_argument("path", type=Path, help="Path to a .xlsx or .csv roster file")
    parser.add_argument("--year", type=int, required=True, help="Integer season year, such as 2025")
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET_NAME,
        help=f"XLSX worksheet name, default: {DEFAULT_SHEET_NAME!r}",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    summary = run_import(args.path, args.year, args.sheet)
    print(json.dumps(summary.model_dump(), indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
