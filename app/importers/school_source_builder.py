from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, cast

from openpyxl import load_workbook  # type: ignore[import-untyped]

from app.schemas.school_sources import (
    SCHOOL_SOURCE_FIELDNAMES,
    SchoolSourceBuildSummary,
    SchoolSourceRow,
)
from app.utils.roster_normalization import clean_text
from app.utils.school_sources import (
    normalize_ipeds_id,
    normalize_school_id,
    normalize_school_source_row,
    school_sort_key,
)

DEFAULT_WORKBOOK_PATH = Path("data/raw/NCAA Roster Data Cleaned.xlsx")
DEFAULT_OUTPUT_PATH = Path("data/schools.csv")


def build_school_source_rows(workbook_path: Path) -> list[SchoolSourceRow]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        id_match_rows = _read_sheet_rows(workbook_path, workbook["ID Match"])
        college_list_rows = _read_sheet_rows(workbook_path, workbook["College List"])
        cleaned_rows = _read_sheet_rows(workbook_path, workbook["Cleaned Data"])
    finally:
        workbook.close()

    canonical_metadata_by_school_id: dict[str, dict[str, str | None]] = {}
    canonical_metadata_by_ipeds_id: dict[str, dict[str, str | None]] = {}
    for row in id_match_rows:
        school_id = normalize_school_id(row.get("School ID"))
        ipeds_id = normalize_ipeds_id(row.get("IPEDS ID"))
        canonical_metadata = {
            "school_id": school_id,
            "ipeds_id": ipeds_id,
            "school_name": clean_text(row.get("School")),
            "city": clean_text(row.get("City")),
            "state": clean_text(row.get("State")),
            "public_private": clean_text(row.get("Public/Private")),
            "conference": clean_text(row.get("Conference")),
            "division": clean_text(row.get("Division")),
        }
        if school_id:
            canonical_metadata_by_school_id[school_id] = canonical_metadata
        if ipeds_id:
            canonical_metadata_by_ipeds_id[ipeds_id] = canonical_metadata

    sources_by_school_id: dict[str, dict[str, str | None]] = {}
    for row in college_list_rows:
        school_id = normalize_school_id(row.get("School ID"))
        if school_id is None:
            continue
        sources_by_school_id[school_id] = {
            "school_id": school_id,
            "school_name": clean_text(row.get("School")),
            "state": clean_text(row.get("State")),
            "division": clean_text(row.get("Division")),
            "roster_url": clean_text(row.get("Roster Page")),
        }

    for row in cleaned_rows:
        school_id = normalize_school_id(row.get("School ID"))
        if school_id is None:
            continue
        source = sources_by_school_id.setdefault(
            school_id,
            {
                "school_id": school_id,
                "school_name": clean_text(row.get("School")),
                "state": clean_text(row.get("State")),
                "division": clean_text(row.get("Division")),
                "roster_url": clean_text(row.get("Roster URL")),
                "ipeds_id": normalize_ipeds_id(row.get("IPEDS ID")),
                "conference": clean_text(row.get("Conference")),
            },
        )
        source.setdefault("roster_url", clean_text(row.get("Roster URL")))
        source.setdefault("ipeds_id", normalize_ipeds_id(row.get("IPEDS ID")))
        source.setdefault("conference", clean_text(row.get("Conference")))

    rows: list[SchoolSourceRow] = []
    seen_ipeds_ids: set[str] = set()
    for school_id, source in sources_by_school_id.items():
        ipeds_id = normalize_ipeds_id(source.get("ipeds_id"))
        metadata: dict[str, str | None] | None = canonical_metadata_by_school_id.get(school_id)
        if metadata is None and ipeds_id is not None:
            metadata = canonical_metadata_by_ipeds_id.get(ipeds_id)

        school_row = normalize_school_source_row(
            SchoolSourceRow(
                school_id=school_id,
                ipeds_id=ipeds_id or (metadata or {}).get("ipeds_id"),
                school_name=(metadata or {}).get("school_name")
                or source.get("school_name")
                or school_id,
                city=(metadata or {}).get("city"),
                state=(metadata or {}).get("state") or source.get("state"),
                public_private=(metadata or {}).get("public_private"),
                division=(metadata or {}).get("division") or source.get("division"),
                conference=(metadata or {}).get("conference") or source.get("conference"),
                roster_url=source.get("roster_url"),
                notes=None,
            )
        )

        if school_row.ipeds_id and school_row.ipeds_id in seen_ipeds_ids:
            continue
        if school_row.ipeds_id:
            seen_ipeds_ids.add(school_row.ipeds_id)
        rows.append(school_row)

    rows.sort(key=school_sort_key)
    return rows


def write_school_sources_csv(
    rows: list[SchoolSourceRow], output_path: Path
) -> SchoolSourceBuildSummary:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as output_file:
        writer = csv.DictWriter(
            output_file,
            fieldnames=SCHOOL_SOURCE_FIELDNAMES,
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(row.model_dump())
    return SchoolSourceBuildSummary(workbook_path=str(output_path), rows_written=len(rows))


def build_school_sources_csv(
    workbook_path: Path = DEFAULT_WORKBOOK_PATH,
    output_path: Path = DEFAULT_OUTPUT_PATH,
) -> SchoolSourceBuildSummary:
    rows = build_school_source_rows(workbook_path)
    summary = write_school_sources_csv(rows, output_path)
    return SchoolSourceBuildSummary(
        workbook_path=str(workbook_path), rows_written=summary.rows_written
    )


def _read_sheet_rows(workbook_path: Path, worksheet: object) -> list[dict[str, object]]:
    del workbook_path
    rows = cast(Any, worksheet).iter_rows(values_only=True)
    headers = [clean_text(header) for header in next(rows)]
    data: list[dict[str, object]] = []
    for row in rows:
        values = {header: value for header, value in zip(headers, row, strict=False) if header}
        if values:
            data.append(values)
    return data
