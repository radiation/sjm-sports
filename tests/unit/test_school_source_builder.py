from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook

from app.importers.school_source_builder import build_school_source_rows, write_school_sources_csv
from app.schemas.school_sources import SchoolSourceRow
from app.utils.school_sources import classify_roster_vendor_from_url, normalize_school_source_row


def test_classify_roster_vendor_from_url_sidearm_patterns() -> None:
    vendor, is_sidearm = classify_roster_vendor_from_url(
        "https://acusports.com/sports/baseball/roster"
    )

    assert vendor == "sidearm"
    assert is_sidearm is True


def test_classify_roster_vendor_from_url_presto_patterns() -> None:
    vendor, is_sidearm = classify_roster_vendor_from_url(
        "https://gobonnies.com/sports/2024/2/1/baseball?path=baseball&portal=prestosports"
    )

    assert vendor == "presto"
    assert is_sidearm is False


def test_normalize_school_source_row_sets_import_enabled() -> None:
    row = normalize_school_source_row(
        SchoolSourceRow(
            school_id="s1",
            school_name=" Example State ",
            roster_url=" https://example.edu/sports/baseball/roster ",
        )
    )

    assert row.school_id == "S1"
    assert row.school_name == "Example State"
    assert row.import_enabled is True
    assert row.roster_vendor == "sidearm"


def test_build_school_source_rows_prefers_id_match_metadata_and_dedupes(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    id_match = workbook.create_sheet("ID Match")
    id_match.append(
        [
            "School ID",
            "IPEDS ID",
            "School",
            "City",
            "State",
            "Public/Private",
            "Conference",
            "Division",
        ]
    )
    id_match.append(
        [
            "S1",
            "1001",
            "Example State University",
            "Example City",
            "TX",
            "Public",
            "Example Conf",
            "NCAA D1",
        ]
    )

    college_list = workbook.create_sheet("College List")
    college_list.append(["School ID", "School", "State", "Division", "Roster Page"])
    college_list.append(
        ["S1", "Example State", "TX", "NCAA D1", "https://example.edu/sports/baseball/roster"]
    )
    college_list.append(
        ["S2", "Second State", "CA", "NCAA D1", "https://second.edu/sports/baseball/roster"]
    )

    cleaned = workbook.create_sheet("Cleaned Data")
    cleaned.append(
        ["School ID", "IPEDS ID", "School", "State", "Division", "Conference", "Roster URL"]
    )
    cleaned.append(
        [
            "S1",
            "1001",
            "Example State Alt",
            "TX",
            "NCAA D1",
            "Other Conf",
            "https://example.edu/sports/baseball/roster",
        ]
    )
    cleaned.append(
        [
            "S2",
            "2002",
            "Second State",
            "CA",
            "NCAA D1",
            "Second Conf",
            "https://second.edu/sports/baseball/roster",
        ]
    )
    cleaned.append(
        [
            "S2",
            "2002",
            "Second State",
            "CA",
            "NCAA D1",
            "Second Conf",
            "https://second.edu/sports/baseball/roster",
        ]
    )

    workbook_path = tmp_path / "schools.xlsx"
    workbook.save(workbook_path)

    rows = build_school_source_rows(workbook_path)

    assert len(rows) == 2
    assert rows[0].school_name == "Example State University"
    assert rows[0].city == "Example City"
    assert rows[1].conference == "Second Conf"


def test_write_school_sources_csv_writes_expected_headers(tmp_path: Path) -> None:
    output_path = tmp_path / "schools.csv"
    summary = write_school_sources_csv(
        [
            SchoolSourceRow(
                school_id="S1",
                school_name="Example State",
                roster_url="https://example.edu/sports/baseball/roster",
                roster_vendor="sidearm",
                is_sidearm=True,
                import_enabled=True,
            )
        ],
        output_path,
    )

    with output_path.open("r", encoding="utf-8", newline="") as generated_file:
        reader = csv.DictReader(generated_file)
        rows = list(reader)

    assert summary.rows_written == 1
    assert rows[0]["school_id"] == "S1"
    assert rows[0]["roster_vendor"] == "sidearm"
